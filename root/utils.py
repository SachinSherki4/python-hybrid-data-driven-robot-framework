import os
from pathlib import Path
from openpyxl import load_workbook
from selenium import webdriver
import random
import string
import yaml
from ntpath import join


class Utils(object):
    def __init__(self):
        """This will get all config data and assign it as class attribute in Key-value pair"""
        with open(Path(__file__).resolve().parent/"config/config.yml", 'r') as file :
            config=yaml.load(file, Loader=yaml.FullLoader)
            file.close()
            for key,value in list(config.items()):
                setattr(self, key, value)
                
                
    def load_excel_data(self):
        # path=os.path.abspath(os.path.dirname(__file__))
        # file_path=os.path.join(path,'data','ParabankData.xlsx')
        excel_path=Path(__file__).resolve().parent/"data"/"ParabankData.xlsx"
        workbook=load_workbook(excel_path)
        sheet=workbook['Registration']
        """This will fetch all headers in one list """
        header=[cell.value for cell in sheet[1]]
        #print(header)
        data=[]
        """This will iterate each row from 2 and assign each value to header in dictionary format"""
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if all(row):
                raw_data=dict(zip(header,row))
                #print(raw_data)
                data.append(raw_data)
        return data   
    
    def get_page_locators(self,locator_page_name):
        """This will fetch respective page locators from config folder based on page name"""
        with open(Path(__file__).resolve().parent/f"config/locators/{locator_page_name}.yml") as file:
            page_locators=yaml.load(file, Loader=yaml.FullLoader)
            file.close()
            return page_locators[locator_page_name]


            """For headless browser testing use Options """
    def launch_browser(self,base_url,browser):
        if browser=="Chrome":
            options=webdriver.ChromeOptions()
            options.add_argument("--headless")
            driver=webdriver.Chrome(options=options)
            #driver=webdriver.Chrome()
            return self.visit_page(driver,base_url)
        elif browser=="Firefox":
            options=webdriver.FirefoxOptions()
            options.add_argument("--headless")
            driver=webdriver.Firefox(options=options)
            #driver=webdriver.Firefox()
            return self.visit_page(driver,base_url)
        elif browser =="Edge":
            options=webdriver.EdgeOptions()
            options.add_argument("--headless")
            driver=webdriver.Edge(options=options)
            #driver=webdriver.Edge()
            return self.visit_page(driver,base_url)
        else:
            print("Please specify the browser in config. file")
    
    def visit_page(self,driver,base_url):
        driver.maximize_window()
        driver.get(base_url)
        return driver
    
    def get_random_strin(self,length):
        random_string=""
        for _ in range(length):
            random_string +="".join(random.choice(string.ascii_letters+string.ascii_letters))
        return random_string
# ut=Utils()
# ##locator=ut.get_page_locators("registration")
# user_data=ut.load_excel_data()
# print(user_data)
# ut.visit_page()
    